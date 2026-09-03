# -*- coding: utf-8 -*-
"""Plantillas por equipo en Excel (una pestana por club) y PDF (una pagina por club).

Sale del mismo catalogo que se cargo en la BD (futbolfantasy), filtrando primer_equipo,
para que lo que se reparte coincida con lo que vera la app. Nacionalidad y edad se
toman de football-data cuando el jugador esta identificado alli. Los cedidos que
futbolfantasy lista en dos clubes se asignan al que le da dorsal, igual que la carga.

Uso:  python tools/plantillas_excel_pdf.py      (necesita exports/carga_inicial_control.csv)
"""
import os, csv, json, datetime, re

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(BASE, 'exports')
CACHE = os.path.join(BASE, 'tools', '.cache')
HOY = datetime.date(2026, 9, 2)
TEMP = '2026/27'

ORDEN = {'PORTERO': 0, 'DEFENSA': 1, 'MEDIO': 2, 'DELANTERO': 3}
ETIQUETA = [('PORTERO', 'Porteros'), ('DEFENSA', 'Defensas'),
            ('MEDIO', 'Centrocampistas'), ('DELANTERO', 'Delanteros')]
COLOR_XL = {'PORTERO': 'F2C14E', 'DEFENSA': '7FB3D5', 'MEDIO': '82C784', 'DELANTERO': 'E57373'}
COLOR_PDF = {'PORTERO': '#B9860B', 'DEFENSA': '#2E6DA4', 'MEDIO': '#2E7D32', 'DELANTERO': '#C62828'}


def edad(fnac):
    if not fnac:
        return ''
    d = datetime.date.fromisoformat(fnac)
    return HOY.year - d.year - ((HOY.month, HOY.day) < (d.month, d.day))


def datos():
    """{equipo: {'corto':..., 'estadio':..., 'jugadores':[...]}} solo primer equipo."""
    fd = json.load(open(os.path.join(CACHE, 'fd_teams.json'), encoding='utf-8'))
    meta, ficha = {}, {}
    for t in fd['teams']:
        meta[t['name']] = {'corto': t.get('shortName') or t['name'], 'estadio': t.get('venue') or ''}
        for p in t.get('squad') or []:
            ficha[(t['name'], p['name'])] = (p.get('nationality') or '', p.get('dateOfBirth') or '')

    # nombre corto de club, el mismo que se cargo en falm.equipo_lfp
    CORTO = {'Deportivo Alavés': 'Alavés', 'Athletic Club': 'Athletic',
             'Club Atlético de Madrid': 'Atlético', 'FC Barcelona': 'Barcelona',
             'Real Betis Balompié': 'Betis', 'RC Celta de Vigo': 'Celta',
             'RC Deportivo La Coruña': 'Deportivo', 'Elche CF': 'Elche',
             'RCD Espanyol de Barcelona': 'Espanyol', 'Getafe CF': 'Getafe',
             'Levante UD': 'Levante', 'Málaga CF': 'Málaga', 'CA Osasuna': 'Osasuna',
             'Real Racing Club de Santander': 'Racing', 'Rayo Vallecano de Madrid': 'Rayo',
             'Real Madrid CF': 'Real Madrid', 'Real Sociedad de Fútbol': 'Real Sociedad',
             'Sevilla FC': 'Sevilla', 'Valencia CF': 'Valencia', 'Villarreal CF': 'Villarreal'}

    # un cedido aparece en dos clubes: se queda en el que le da dorsal
    todas = list(csv.DictReader(open(os.path.join(OUT, 'carga_inicial_control.csv'),
                                     encoding='utf-8-sig'), delimiter=';'))
    por_slug = {}
    for r in todas:
        otra = por_slug.get(r['slug'])
        if otra is None or (not otra['dorsal'] and r['dorsal']):
            por_slug[r['slug']] = r
    # primer equipo si lo es en cualquiera de sus fichas, igual que en la BD
    for r in todas:
        if r['primer_equipo'] == 'SI':
            por_slug[r['slug']]['primer_equipo'] = 'SI'

    equipos = {}
    for r in por_slug.values():
            if r['primer_equipo'] != 'SI':
                continue
            nac, fnac = ficha.get((r['equipo'], r['nombre_largo_football_data']), ('', ''))
            club = CORTO.get(r['equipo'], r['equipo'])
            e = equipos.setdefault(club, dict(meta.get(r['equipo'], {'corto': club, 'estadio': ''}),
                                              corto=club, jugadores=[]))
            e['jugadores'].append({
                'dorsal': int(r['dorsal']) if r['dorsal'] else None,
                'nombre': r['jugador'], 'posicion': r['posicion'],
                'nac': nac, 'edad': edad(fnac),
            })
    for e in equipos.values():
        e['jugadores'].sort(key=lambda j: (ORDEN[j['posicion']], j['dorsal'] is None,
                                           j['dorsal'] or 0, j['nombre']))
    return dict(sorted(equipos.items()))


def excel(equipos, ruta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    s = Side(style='thin', color='D9D9D9')
    thin = Border(left=s, right=s, top=s, bottom=s)
    usados = set()
    for nombre, e in equipos.items():
        hoja = re.sub(r'[\[\]:*?/\\]', '', e['corto'])[:31].strip()
        base, i = hoja, 2
        while hoja.lower() in usados:
            hoja, i = '%s %d' % (base[:28], i), i + 1
        usados.add(hoja.lower())

        ws = wb.create_sheet(hoja)
        ws.append([nombre])
        ws['A1'].font = Font(bold=True, size=13, color='1F3864')
        ws.append(['%s  |  %d jugadores  |  LaLiga %s' % (e['estadio'] or '-', len(e['jugadores']), TEMP)])
        ws.append([])
        ws.append(['Dorsal', 'Jugador', 'Posicion', 'Nacionalidad', 'Edad'])
        for c in ws[ws.max_row]:
            c.fill = PatternFill('solid', fgColor='1F3864')
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin
        for j in e['jugadores']:
            ws.append([j['dorsal'] or '', j['nombre'], j['posicion'], j['nac'], j['edad']])
            for col in range(1, 6):
                ws.cell(ws.max_row, col).border = thin
            ws.cell(ws.max_row, 1).alignment = Alignment(horizontal='center')
            ws.cell(ws.max_row, 3).fill = PatternFill('solid', fgColor=COLOR_XL[j['posicion']])
        for col, ancho in enumerate([8, 32, 14, 20, 8], 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
        ws.freeze_panes = 'A5'
    wb.save(ruta)


def pdf(equipos, ruta):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    doc = SimpleDocTemplate(ruta, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm,
                            title='Plantillas LaLiga ' + TEMP, author='FALM')
    ss = getSampleStyleSheet()
    H1 = ParagraphStyle('H1', parent=ss['Heading1'], fontSize=17,
                        textColor=colors.HexColor('#1F3864'), spaceAfter=1)
    SUB = ParagraphStyle('SUB', parent=ss['Normal'], fontSize=8.5,
                         textColor=colors.HexColor('#666666'), spaceAfter=6)
    BASE_TS = [('FONTSIZE', (0, 0), (-1, -1), 8.5),
               ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
               ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
               ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#DDDDDD')),
               ('BOTTOMPADDING', (0, 0), (-1, -1), 3), ('TOPPADDING', (0, 0), (-1, -1), 3)]
    story = []

    total = sum(len(e['jugadores']) for e in equipos.values())
    story.append(Paragraph('Plantillas LaLiga ' + TEMP, H1))
    story.append(Paragraph('Mercado cerrado &middot; %d equipos, %d jugadores &middot; %s'
                           % (len(equipos), total, HOY.strftime('%d/%m/%Y')), SUB))
    res = [['Equipo', 'Tot', 'POR', 'DEF', 'MED', 'DEL']]
    for nombre, e in equipos.items():
        c = {k: sum(1 for j in e['jugadores'] if j['posicion'] == k) for k in ORDEN}
        res.append([nombre, len(e['jugadores']), c['PORTERO'], c['DEFENSA'], c['MEDIO'], c['DELANTERO']])
    res.append(['TOTAL', total] + [sum(1 for e in equipos.values() for j in e['jugadores']
                                       if j['posicion'] == k) for k in ['PORTERO', 'DEFENSA', 'MEDIO', 'DELANTERO']])
    t = Table(res, colWidths=[80*mm, 18*mm, 18*mm, 18*mm, 18*mm, 18*mm], repeatRows=1)
    t.setStyle(TableStyle(BASE_TS + [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F3864')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F2F5FA')])]))
    story += [t, PageBreak()]

    for nombre, e in equipos.items():
        story.append(Paragraph(nombre, H1))
        story.append(Paragraph('%s &middot; %d jugadores' % (e['estadio'] or '-', len(e['jugadores'])), SUB))
        for cod, etq in ETIQUETA:
            grupo = [j for j in e['jugadores'] if j['posicion'] == cod]
            if not grupo:
                continue
            data = [['#', '%s  (%s)' % (etq, cod), 'Nacionalidad', 'Edad']]
            for j in grupo:
                data.append([j['dorsal'] or '', j['nombre'], j['nac'], j['edad']])
            tt = Table(data, colWidths=[14*mm, 90*mm, 50*mm, 26*mm], repeatRows=1)
            tt.setStyle(TableStyle(BASE_TS + [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(COLOR_PDF[cod])),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F7F7')])]))
            story += [tt, Spacer(1, 3*mm)]
        story.append(PageBreak())
    doc.build(story[:-1])


def main():
    equipos = datos()
    x = os.path.join(OUT, 'plantillas_laliga_2026-27.xlsx')
    p = os.path.join(OUT, 'plantillas_laliga_2026-27.pdf')
    excel(equipos, x)
    pdf(equipos, p)
    total = sum(len(e['jugadores']) for e in equipos.values())
    sin_nac = sum(1 for e in equipos.values() for j in e['jugadores'] if not j['nac'])
    print('XLSX:', x)
    print('PDF :', p)
    print('equipos: %d | jugadores: %d | sin nacionalidad/edad: %d' % (len(equipos), total, sin_nac))


if __name__ == '__main__':
    main()
